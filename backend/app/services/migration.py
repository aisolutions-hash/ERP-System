"""Excel -> PostgreSQL migration for the Kalika ERP.

Reads the source workbook (RAW  MATERIAL, STORE, PRODUCTION, DISPATCH, ORDER
sheets) and populates the database following the agreed schema mapping.
Handles per-sheet quirks: section rows, G.TOTAL/SUB-TOTAL rows, stale daily
date columns, and off-by-one column alignments in some sheets.
"""
import re
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from ..database import Base, SessionLocal
from ..models import (
    Customer, Dispatch, DispatchLine, DispatchStatus, Inventory, MigrationLog,
    MovementType, OrderStatus, Plant, Product, ProductCategory, ProductionMovement,
    ProductionOrder, ProductionStatus, PurchaseOrder, PurchaseOrderLine,
    PurchaseStatus, RawMaterialBalance, SalesOrder, SalesOrderLine,
    StockMovement,
)

SKIP_LABELS = {"G.TOTAL", "SUB-TOTAL", "TOTAL", "GRAND TOTAL", "GRAND-TOTAL"}
PLANT_NAMES = {"G-6", "G-06", "C-33", "PHP", "CPG", "CIL-085", "DAIMLER",
               "IPDC", "CFS", "TCL", "TCL - 2 -3", "TCL-2-3", "EATON",
               "RECORN", "JAKSON", "jakson", "JACKSON"}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_PERSON_RE = re.compile(r"^[A-Z\s/]+$")


def _norm(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).replace("\t", " ").replace("\n", " ")).strip()


def _num(v):
    try:
        f = float(v)
        return f
    except (TypeError, ValueError):
        return None


def _daily_date_cols(header):
    return [i for i, c in enumerate(header) if _DATE_RE.match(c.strip())]


def _find_header_row(ws):
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [_norm(c).upper() for c in row]
        if any(c in ("ITEM CODE", "ITEM CODE ") for c in cells):
            return i
    return None


def _parse_date(v):
    if v is None:
        return None
    if isinstance(v, date):
        return v
    try:
        from openpyxl.utils.datetime import from_excel
        return from_excel(float(v)).date()
    except Exception:
        try:
            return date.fromisoformat(str(v)[:10])
        except Exception:
            return None


def _get_or_create_product(db, item_code, model, category, uom="kg", source=""):
    ic, m = _norm(item_code), _norm(model)
    p = db.query(Product).filter(Product.item_code == ic, Product.model == m).first()
    if p is None:
        p = Product(item_code=ic, model=m or ic or "(unnamed)", category=category,
                    uom=uom, source_excel=source)
        db.add(p)
        db.flush()
    return p


def _get_or_create_plant(db, name, customer_id=None, source=""):
    name = _norm(name)
    if not name:
        return None
    pl = db.query(Plant).filter(Plant.name == name).first()
    if pl is None:
        pl = Plant(name=name, customer_id=customer_id, source_excel=source)
        db.add(pl)
        db.flush()
    return pl


def _get_or_create_customer(db, name, is_plant=False, source=""):
    name = _norm(name)
    if not name:
        return None
    c = db.query(Customer).filter(Customer.name == name).first()
    if c is None:
        c = Customer(name=name, is_plant=is_plant, source_excel=source)
        db.add(c)
        db.flush()
    return c


def _inventory_for(db, product, plant=None):
    inv = db.query(Inventory).filter(
        Inventory.product_id == product.id,
        Inventory.plant_id == (plant.id if plant else None),
    ).first()
    if inv is None:
        inv = Inventory(product_id=product.id, plant_id=plant.id if plant else None)
        db.add(inv)
        db.flush()
    return inv


def _add_movement(db, product, mtype, qty, txn_date, source, remarks="", plant=None):
    if qty in (None, 0):
        return
    db.add(StockMovement(product_id=product.id, plant_id=plant.id if plant else None,
                         movement_type=mtype, quantity=qty, transaction_date=txn_date,
                         source_excel=source, remarks=remarks))


def _lookup(row, idx, name):
    i = idx.get(name)
    if i is None or i >= len(row):
        return None
    return row[i]


def _skip_row(model):
    m = _norm(model).upper()
    return m in SKIP_LABELS or m.startswith("G. TOTAL") or m.startswith("SUB-TOTAL")


# ---------------------------------------------------------------------------
# Sheet migrations
# ---------------------------------------------------------------------------

def migrate_raw_material(db, ws, source, report_date):
    hr = _find_header_row(ws)
    if not hr:
        return 0, 0
    header = [_norm(c) for c in next(ws.iter_rows(min_row=hr, max_row=hr, values_only=True))]
    idx = {name: i for i, name in enumerate(header) if name}
    daily = _daily_date_cols(header)
    imported = skipped = 0
    for row in ws.iter_rows(min_row=hr + 1, values_only=True):
        if not any(row):
            continue
        model = _norm(_lookup(row, idx, "MODEL")) or _norm(_lookup(row, idx, "ITEM CODE"))
        if not model:
            skipped += 1
            continue
        if _skip_row(model):
            skipped += 1
            continue
        prod = _get_or_create_product(db, _lookup(row, idx, "ITEM CODE"), model,
                                      ProductCategory.raw_material, uom="kg", source=source)
        inward = _num(_lookup(row, idx, "INWARD QTY"))
        db.add(RawMaterialBalance(
            product_id=prod.id, report_date=report_date,
            schedule_qty=_num(_lookup(row, idx, "SCHEDULE")),
            ask_till_date=_num(_lookup(row, idx, "ASK TILL DATE")),
            inward_qty=inward,
            completion_pct=_num(_lookup(row, idx, "% COMP")),
            balance_qty=_num(_lookup(row, idx, "BALANCE QTY")),
            opening_stock=_num(_lookup(row, idx, "OPNING STOCK")),
            notes=source,
        ))
        for di in daily:
            v = _num(row[di]) if di < len(row) else None
            if v:
                _add_movement(db, prod, MovementType.receipt, v, date.fromisoformat(header[di]),
                              source, remarks="Raw material inward")
        imported += 1
    return imported, skipped


def migrate_store(db, ws, source, report_date):
    hr = _find_header_row(ws)
    if not hr:
        return 0, 0
    header = [_norm(c) for c in next(ws.iter_rows(min_row=hr, max_row=hr, values_only=True))]
    idx = {name: i for i, name in enumerate(header) if name}
    current_plant = None
    imported = skipped = 0
    for row in ws.iter_rows(min_row=hr + 1, values_only=True):
        if not any(row):
            continue
        # section header: col0 empty, col1 empty, col2 = plant name (uppercase / short)
        c0, c1, c2 = _norm(row[0]), _norm(row[1]), _norm(row[2])
        if not c0 and not c1 and c2 and (c2.upper() in {p.upper() for p in PLANT_NAMES} or _num(row[2]) is None and len(c2) < 40 and any(ch.isdigit() for ch in c2) is False and c2 == c2.upper()):
            current_plant = _get_or_create_plant(db, c2, source=source)
            skipped += 1
            continue
        item_code = _lookup(row, idx, "ITEM CODE")
        model = _lookup(row, idx, "MODEL")
        if not _norm(model) and not _norm(item_code):
            skipped += 1
            continue
        if _skip_row(_norm(model) or _norm(item_code)):
            skipped += 1
            continue
        prod = _get_or_create_product(db, item_code, model, ProductCategory.store,
                                      uom="kg", source=source)
        po_no = _norm(_lookup(row, idx, "PO NO"))
        if po_no:
            po = db.query(PurchaseOrder).filter(PurchaseOrder.po_number == po_no).first()
            if po is None:
                po = PurchaseOrder(po_number=po_no, order_date=report_date,
                                   status=PurchaseStatus.ordered, source_excel=source)
                db.add(po)
                db.flush()
            if not any(l.product_id == prod.id for l in po.lines):
                po.lines.append(PurchaseOrderLine(
                    product_id=prod.id, description=f"{po_no} / {prod.model}",
                    quantity=_num(_lookup(row, idx, "PO QTY")) or 0, rate=None, amount=None))
        inv = _inventory_for(db, prod, current_plant)
        opening = _num(_lookup(row, idx, "OPNING STOCK")) or 0
        outward = _num(_lookup(row, idx, "OUTWARD")) or 0
        balance = _num(_lookup(row, idx, "BALANCE QTY"))
        inv.opening_stock += opening
        inv.issued_qty += outward
        # source-of-truth balance is the reported BALANCE QTY when present
        inv.current_stock += balance if balance is not None else opening - outward
        if outward:
            _add_movement(db, prod, MovementType.issue, outward, report_date, source,
                          remarks="Store outward", plant=current_plant)
        imported += 1
    return imported, skipped


def migrate_production(db, ws, source, report_date):
    hr = _find_header_row(ws)
    if not hr:
        return 0, 0
    header = [_norm(c) for c in next(ws.iter_rows(min_row=hr, max_row=hr, values_only=True))]
    idx = {name: i for i, name in enumerate(header) if name}
    daily = _daily_date_cols(header)
    imported = skipped = 0
    for row in ws.iter_rows(min_row=hr + 1, values_only=True):
        if not any(row):
            continue
        model = _norm(_lookup(row, idx, "MODEL")) or _norm(_lookup(row, idx, "ITEM CODE"))
        if not model:
            skipped += 1
            continue
        if _skip_row(model):
            skipped += 1
            continue
        prod = _get_or_create_product(db, _lookup(row, idx, "ITEM CODE"), model,
                                      ProductCategory.finished, uom="kg", source=source)
        produced = _num(_lookup(row, idx, "PRODUCTION QTY")) or 0
        schedule = _num(_lookup(row, idx, "SCHEDULE")) or 0
        po = ProductionOrder(
            order_no=f"MIG-PR-{source.split('/')[-1]}-{imported + 1:03d}",
            product_id=prod.id,
            schedule_qty=schedule,
            ask_till_date=_num(_lookup(row, idx, "ASK TILL DATE")),
            produced_qty=produced,
            completion_pct=_num(_lookup(row, idx, "% COMP")),
            balance_qty=_num(_lookup(row, idx, "BALANCE QTY")) if _lookup(row, idx, "BALANCE QTY") is not None else schedule - produced,
            opening_stock=_num(_lookup(row, idx, "OPNING STOCK")) or 0,
            status=ProductionStatus.completed if produced >= schedule and schedule > 0
                   else (ProductionStatus.in_production if produced > 0 else ProductionStatus.planned),
            report_date=report_date, source_excel=source,
        )
        db.add(po)
        db.flush()
        for di in daily:
            v = _num(row[di]) if di < len(row) else None
            if v:
                db.add(ProductionMovement(production_order_id=po.id, quantity=v,
                                          production_date=date.fromisoformat(header[di]),
                                          source_excel=source))
                _add_movement(db, prod, MovementType.production_output, v,
                              date.fromisoformat(header[di]), source, remarks="Production output")
        if produced:
            inv = _inventory_for(db, prod)
            inv.received_qty += produced
            inv.current_stock += produced
        imported += 1
    return imported, skipped


def migrate_dispatch(db, ws, source, report_date):
    hr = _find_header_row(ws)
    if not hr:
        return 0, 0
    header = [_norm(c) for c in next(ws.iter_rows(min_row=hr, max_row=hr, values_only=True))]
    idx = {name: i for i, name in enumerate(header) if name}
    current_customer = None
    current_plant = None
    imported = skipped = 0
    for row in ws.iter_rows(min_row=hr + 1, values_only=True):
        if not any(row):
            continue
        c0, c1 = _norm(row[0]), _norm(row[1])
        # section: two text cells (salesperson + plant) followed by a number
        if c0 and c1 and _num(c0) is None and _num(c1) is None and _PERSON_RE.match(c0) \
                and not (len(c0) > 6 and any(ch in c0 for ch in "0123456789.")):
            cust = _get_or_create_customer(db, f"{c0}/{c1}", is_plant=True, source=source)
            current_customer = cust
            current_plant = _get_or_create_plant(db, c1, customer_id=cust.id if cust else None, source=source)
            skipped += 1
            continue
        item_code = _lookup(row, idx, "ITEM CODE")
        model = _lookup(row, idx, "MODEL")
        if not _norm(model) and not _norm(item_code):
            skipped += 1
            continue
        if _skip_row(_norm(model) or _norm(item_code)):
            skipped += 1
            continue
        prod = _get_or_create_product(db, item_code, model, ProductCategory.finished,
                                      uom="kg", source=source)
        sched = _num(_lookup(row, idx, "SCHEDULE")) or 0
        dispatched = _num(_lookup(row, idx, "DISPATCH"))
        dispatched = dispatched if dispatched is not None else 0
        disp = Dispatch(
            dispatch_no=f"MIG-DP-{source.split('/')[-1]}-{imported + 1:03d}",
            customer_id=current_customer.id if current_customer else None,
            plant_id=current_plant.id if current_plant else None,
            sales_person=current_customer.name.split("/")[0] if current_customer else "",
            schedule_qty=sched,
            ask_till_date=_num(_lookup(row, idx, "ASK TILL DATE")),
            dispatched_qty=dispatched,
            completion_pct=_num(_lookup(row, idx, "% COMP")),
            balance_qty=_num(_lookup(row, idx, "BALANCE QTY")) if _lookup(row, idx, "BALANCE QTY") is not None else sched - dispatched,
            opening_stock=_num(_lookup(row, idx, "OPNING STOCK")) or 0,
            status=DispatchStatus.completed if dispatched >= sched and sched > 0
                   else (DispatchStatus.partial if dispatched > 0 else DispatchStatus.pending),
            report_date=report_date, source_excel=source,
        )
        db.add(disp)
        db.flush()
        db.add(DispatchLine(dispatch_id=disp.id, product_id=prod.id, quantity=dispatched,
                            dispatch_date=report_date, source_excel=source))
        if dispatched:
            _add_movement(db, prod, MovementType.dispatch, dispatched, report_date, source,
                          remarks="Dispatch", plant=current_plant)
            inv = _inventory_for(db, prod, current_plant)
            inv.issued_qty += dispatched
            inv.current_stock -= dispatched
        imported += 1
    return imported, skipped


def migrate_orders(db, ws, source, report_date):
    """ORDER sheet has a two-row header: DATE|CUSTOMER ORDER then MODEL|QTY|CUSTOMER|REMARK."""
    rows = list(ws.iter_rows(values_only=True))
    # find the header pair: a row containing DATE / CUSTOMER ORDER followed by MODEL/QTY
    hr = None
    for i in range(len(rows) - 1):
        top = [_norm(c).upper() for c in rows[i]]
        bottom = [_norm(c).upper() for c in rows[i + 1]]
        if "DATE" in top and "MODEL" in bottom:
            hr = i + 2  # 1-based row of the MODEL header row (skip it when reading data)
            break
    if hr is None:
        return 0, 0
    idx = {"DATE": 0, "MODEL": 1, "QTY": 2, "CUSTOMER": 3, "REMARK": 4}
    imported = skipped = 0
    for row in ws.iter_rows(min_row=hr + 1, values_only=True):
        if not any(row):
            continue
        model = _norm(_lookup(row, idx, "MODEL"))
        cust_name = _norm(_lookup(row, idx, "CUSTOMER"))
        if not model or not cust_name:
            skipped += 1
            continue
        # strip trailing size codes glued onto model (e.g. '500 X 51 X 6\t200')
        clean_model = re.split(r"[\t\n]", model)[0].strip()
        if not clean_model:
            skipped += 1
            continue
        cust = _get_or_create_customer(db, cust_name, source=source)
        prod = _get_or_create_product(db, "", clean_model, ProductCategory.finished,
                                      uom="kg", source=source)
        odate = _parse_date(_lookup(row, idx, "DATE")) or report_date
        qty = _num(_lookup(row, idx, "QTY"))
        o = SalesOrder(
            order_no=f"MIG-SO-{source.split('/')[-1]}-{imported + 1:03d}",
            customer_id=cust.id if cust else None,
            order_date=odate,
            status=OrderStatus.new, source_excel=source,
        )
        o.lines.append(SalesOrderLine(product_id=prod.id, description=clean_model,
                                      quantity=qty or 0,
                                      unit_price=None, amount=None))
        db.add(o)
        imported += 1
    return imported, skipped


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

def _find_sheet(wb, name):
    target = re.sub(r"\s+", " ", name).lower()
    for title in wb.sheetnames:
        if re.sub(r"\s+", " ", title).lower() == target:
            return title
    for title in wb.sheetnames:
        if re.sub(r"\s+", " ", title).lower().startswith(target[:6]):
            return title
    return None


SHEET_MAPPING = [
    ("RAW  MATERIAL", migrate_raw_material),
    ("STORE", migrate_store),
    ("PRODUCTION", migrate_production),
    ("DISPATCH", migrate_dispatch),
    ("ORDER", migrate_orders),
]


def _already_migrated(db, sheet_name, xlsx_path, report_date):
    """True if this sheet was already imported for this report date."""
    logs = db.query(MigrationLog).filter(
        MigrationLog.source_sheet == sheet_name,
        MigrationLog.status == "completed",
    ).all()
    for log in logs:
        if f"report_date={report_date}" in log.assumptions:
            return True
    return False


def run_migration(xlsx_path: Path, report_date: date | None = None, reset: bool = False) -> dict:
    report_date = report_date or date.today()
    db = SessionLocal()
    try:
        if reset:
            for t in reversed(Base.metadata.sorted_tables):
                db.execute(t.delete())
            db.commit()
        wb = load_workbook(xlsx_path, data_only=True)
        results = []
        for sheet_name, fn in SHEET_MAPPING:
            actual = _find_sheet(wb, sheet_name)
            if actual is None:
                results.append({"sheet": sheet_name, "imported": 0, "skipped": 0, "error": "sheet not found"})
                continue
            if not reset and _already_migrated(db, actual, xlsx_path, report_date):
                results.append({"sheet": actual, "imported": 0, "skipped": 0, "skipped_existing": True})
                continue
            try:
                imported, skipped = fn(db, wb[actual], actual, report_date)
                db.commit()
                db.add(MigrationLog(source_sheet=actual, target_table=fn.__name__,
                                    rows_imported=imported, rows_skipped=skipped,
                                    assumptions=f"report_date={report_date} source={xlsx_path}"))
                db.commit()
                results.append({"sheet": actual, "imported": imported, "skipped": skipped})
            except Exception as e:  # noqa: BLE001
                db.rollback()
                db.add(MigrationLog(source_sheet=actual, target_table=fn.__name__,
                                    rows_imported=0, rows_skipped=0,
                                    assumptions=f"error: {e} source={xlsx_path}", status="failed"))
                db.commit()
                results.append({"sheet": actual, "imported": 0, "skipped": 0, "error": str(e)})
        return {"results": results}
    finally:
        db.close()