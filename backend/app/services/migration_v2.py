"""Normalized Excel -> ERP migration (v2).

Safe path: imports the parsed workbook into an ISOLATED staging database
(import batch). The live PostgreSQL database is never touched by this module.

Business rules applied (confirmed decisions):
* Reporting period = August 2026 (stale 2025 date headers preserved as source,
  remapped into the confirmed period, flagged as warnings).
* TCL-2 and TCL-3 are separate customers; historical 'TCL-2-3'/'TCL - 2 -3'
  sections are preserved as-is, flagged NEEDS_BUSINESS_CONFIRMATION.
* STORE 'PO NO' is kept as TEXT customer/order-side reference on the order
  line; no supplier purchase orders are created from it.
* Negative balances are preserved (over-dispatch is valid business data).
* Confirmed manufactured families stay MANUFACTURED; MIXED only when the item
  also has trading evidence (STORE row carrying a PO NO).
* Excel derived 'OPNING STOCK' (=BALANCE/BalanceDays) is a planning metric and
  is NEVER imported as physical inventory.
"""
from __future__ import annotations

import datetime as dt
import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

from ..models import (
    ConfirmationStatus, Customer, CustomerAlias, Dispatch, DispatchLine,
    DispatchStatus, ImportBatch, MovementType, OrderStatus, OrderType, Plan,
    PlanType, Product, ProductAlias, ProductCategory, ProductSourceType,
    ProductionMovement, ProductionOrder, ProductionStatus, RawMaterialBalance,
    ReportingPeriod, SalesOrder, SalesOrderLine, Salesperson, StockMovement,
)
from .excel_parser_v2 import (
    ParsedRow, ParsedSection, SheetParse, norm_num, norm_text,
    parse_report_sheet,
)

PERIOD_YEAR, PERIOD_MONTH = 2026, 8
PERIOD_LABEL = f"{PERIOD_YEAR}-{PERIOD_MONTH:02d}"

# ---------------------------------------------------------------------------
# classification tables
# ---------------------------------------------------------------------------

MANUFACTURED_FAMILIES = [
    "STRETCH FILM", "POLYBAG", "POLY BOX", "VCI BAG & BOX",
    "BOPP / PACKAGING TAPE", "MASKING TAPE", "TARPAULINE", "HDPE BAG",
    "SHRINK FILM", "HAND GLOVES", "STRAP & BUCKLES",
]


def product_family(model: str) -> str:
    m = model.upper()
    if "STRETCH" in m or "STRECH" in m:
        return "STRETCH FILM"
    if "VCI" in m:
        return "VCI BAG & BOX"
    if "POLY BOX" in m or "KIT BOX" in m or "BOX BAG" in m:
        return "POLY BOX"
    if "BOPP" in m or "CELLO TAPE" in m:
        return "BOPP / PACKAGING TAPE"
    if "MASKING" in m:
        return "MASKING TAPE"
    if "TARPAULIN" in m or "TARPULINE" in m or "TARPULIN" in m or "TARPAULINE" in m:
        return "TARPAULINE"
    if "HDPE" in m:
        return "HDPE BAG"
    if "SHRINK" in m:
        return "SHRINK FILM"
    if "GLOVES" in m or "HAND GLOVES" in m:
        return "HAND GLOVES"
    if "STRAP" in m or "BUCKLE" in m:
        return "STRAP & BUCKLES"
    if "POLY" in m or "POLYBAG" in m or "POLLY BAG" in m or "POLYTHENE" in m or "LDPE" in m:
        return "POLYBAG"
    return ""


def canonical_customer(name: str) -> str:
    n = norm_text(name)
    u = n.upper()
    if u in ("JAKSON", "JACKSON"):
        return "JAKSON"
    if u in ("G-6", "G-06", "G6"):
        return "G-06"
    if u in ("TCL - 2 -3", "TCL-2-3", "TCL-2 -3", "TCL -2-3"):
        return "TCL-2-3"
    return n


# ---------------------------------------------------------------------------
# context
# ---------------------------------------------------------------------------

class ImportContext:
    def __init__(self, db, batch: ImportBatch, period: ReportingPeriod, source_file: str):
        self.db = db
        self.batch = batch
        self.period = period
        self.source_file = source_file
        self.customers: dict[str, Customer] = {}
        self.salespersons: dict[str, Salesperson] = {}
        self.products: dict[tuple[str, str], Product] = {}
        self.aliases_created = 0
        self.evidence: dict[tuple[str, str], set] = defaultdict(set)  # product key -> sheets with activity
        self.trading_evidence: set = set()    # keys with STORE PO NO
        self.production_evidence: set = set() # keys with production daily output
        self.warnings: list[str] = []
        self.stats: dict[str, int] = defaultdict(int)

    # -- masters ------------------------------------------------------
    def customer(self, raw_name: str, *, status: str = "CONFIRMED",
                 source_sheet: str = "", source_row: int | None = None,
                 remarks: str = "") -> Customer | None:
        canon = canonical_customer(raw_name)
        if not canon:
            return None
        # centralized confirmation rules (confirmed business decisions)
        if canon == "TCL-2-3" and not remarks:
            status = "NEEDS_BUSINESS_CONFIRMATION"
            remarks = ("Historical combined TCL-2-3 section; TCL-2/TCL-3 split "
                       "not determinable from source")
        if canon == "OTHER" and not remarks:
            status = "NEEDS_BUSINESS_CONFIRMATION"
            remarks = "Ambiguous section label"
        c = self.customers.get(canon)
        if c is None:
            c = Customer(name=canon, source_excel=source_sheet,
                         confirmation_status=status, notes=remarks)
            self.db.add(c)
            self.db.flush()
            self.customers[canon] = c
            self.stats["customers"] += 1
        elif status == "NEEDS_BUSINESS_CONFIRMATION" and c.confirmation_status != status:
            c.confirmation_status = status
            c.notes = remarks or c.notes
        if norm_text(raw_name) != canon:
            self.db.add(CustomerAlias(customer_id=c.id, alias=norm_text(raw_name),
                                      match_rule="SOURCE_VARIANT", source_sheet=source_sheet))
            self.aliases_created += 1
        return c

    def salesperson(self, name: str) -> Salesperson | None:
        n = norm_text(name)
        if not n:
            return None
        s = self.salespersons.get(n)
        if s is None:
            s = Salesperson(name=n)
            self.db.add(s)
            self.db.flush()
            self.salespersons[n] = s
            self.stats["salespersons"] += 1
        return s

    def product(self, code: str, model: str, *, category: ProductCategory,
                source_sheet: str, source_row: int | None) -> Product:
        ic, m = norm_text(code), norm_text(model)
        key = (ic, m or ic or "(unnamed)")
        p = self.products.get(key)
        if p is None:
            p = Product(item_code=ic, model=key[1], category=category,
                        uom="kg", family=product_family(key[1]),
                        source_excel=source_sheet, source_sheet=source_sheet,
                        source_row=source_row, import_batch_id=self.batch.id)
            self.db.add(p)
            self.db.flush()
            self.products[key] = p
            self.stats["products"] += 1
        return p


# ---------------------------------------------------------------------------
# sheet importers
# ---------------------------------------------------------------------------

SALES_STATUSES_DONE = {"DONE"}


def import_raw(ctx: ImportContext, p: SheetParse):
    for sec in p.sections:
        for r in sec.rows:
            model = norm_text(r.values.get("MODEL"))
            if not model:
                continue
            prod = ctx.product("", model, category=ProductCategory.raw_material,
                               source_sheet=p.sheet, source_row=r.row)
            ctx.db.add(RawMaterialBalance(
                product_id=prod.id, report_date=dt.date(PERIOD_YEAR, PERIOD_MONTH, 31),
                schedule_qty=norm_num(r.values.get("SCHEDULE")),
                ask_till_date=norm_num(r.values.get("ASK TILL DATE")),
                inward_qty=norm_num(r.values.get("INWARD QTY")),
                completion_pct=norm_num(r.values.get("% COMP")),
                balance_qty=norm_num(r.values.get("BALANCE QTY")),
                opening_stock=norm_num(r.values.get("OPNING STOCK")),
                source_row=r.row, import_batch_id=ctx.batch.id, notes=p.sheet))
            ctx.stats["raw_material_balances"] += 1
            opening = norm_num(r.values.get("OPNING STOCK"))
            if opening:
                ctx.db.add(StockMovement(
                    product_id=prod.id, movement_type=MovementType.opening,
                    quantity=opening, transaction_date=dt.date(PERIOD_YEAR, PERIOD_MONTH, 1),
                    ref_type="raw_material_balance", source_excel=p.sheet,
                    source_sheet=p.sheet, source_row=r.row,
                    import_batch_id=ctx.batch.id, remarks="Raw material opening stock"))
                ctx.stats["stock_movements"] += 1
            for d, q in r.daily.items():
                ctx.db.add(StockMovement(
                    product_id=prod.id, movement_type=MovementType.purchase_receipt,
                    quantity=q, transaction_date=d, ref_type="raw_material_balance",
                    source_excel=p.sheet, source_sheet=p.sheet, source_row=r.row,
                    import_batch_id=ctx.batch.id, remarks="Raw material inward"))
                ctx.stats["stock_movements"] += 1
            ctx.evidence[(prod.item_code, prod.model)].add("RAW")


def import_store(ctx: ImportContext, p: SheetParse):
    for sec in p.sections:
        cust = None
        if sec.label:
            cust = ctx.customer(sec.label, source_sheet=p.sheet,
                                source_row=sec.subtotal_row)
        else:
            ctx.warnings.append(
                f"STORE rows {[r.row for r in sec.rows][:5]}...: no customer section label "
                f"(subtotal row {sec.subtotal_row}); customer left NULL (NEEDS_BUSINESS_CONFIRMATION)")
        for r in sec.rows:
            code = r.code
            model = norm_text(r.values.get("MODEL"))
            if not model and not code:
                continue
            po_no = norm_text(r.values.get("PO NO"))
            prod = ctx.product(code, model, category=ProductCategory.store,
                               source_sheet=p.sheet, source_row=r.row)
            key = (prod.item_code, prod.model)
            ctx.evidence[key].add("STORE")
            if po_no:
                ctx.trading_evidence.add(key)
            qty = norm_num(r.values.get("PO QTY")) or 0
            order = SalesOrder(
                order_no=f"SO-TRD-{PERIOD_LABEL}-{ctx.stats['sales_orders'] + 1:04d}",
                customer_id=cust.id if cust else None,
                order_type=OrderType.trading, period_id=ctx.period.id,
                order_date=dt.date(PERIOD_YEAR, PERIOD_MONTH, 1),
                status=OrderStatus.confirmed,
                remarks="Imported from STORE sheet; PO NO semantics UNKNOWN (customer/order-side reference)",
                source_excel=p.sheet, source_sheet=p.sheet, source_row=r.row,
                import_batch_id=ctx.batch.id)
            order.lines.append(SalesOrderLine(
                product_id=prod.id, description=prod.model, quantity=qty,
                customer_po_no=po_no, source_row=r.row, import_batch_id=ctx.batch.id))
            ctx.db.add(order)
            ctx.db.flush()
            ctx.stats["sales_orders"] += 1
            ctx.stats["sales_order_lines"] += 1
            for d, q in r.daily.items():
                ctx.db.add(StockMovement(
                    product_id=prod.id, movement_type=MovementType.dispatch,
                    quantity=q, transaction_date=d, ref_type="sales_order", ref_id=order.id,
                    source_excel=p.sheet, source_sheet=p.sheet, source_row=r.row,
                    import_batch_id=ctx.batch.id,
                    remarks=f"STORE outward ({sec.label or 'unassigned'})"))
                ctx.stats["stock_movements"] += 1


def import_production(ctx: ImportContext, p: SheetParse):
    for sec in p.sections:
        for r in sec.rows:
            code = r.code
            model = norm_text(r.values.get("MODEL"))
            if not model and not code:
                continue
            prod = ctx.product(code, model, category=ProductCategory.finished,
                               source_sheet=p.sheet, source_row=r.row)
            key = (prod.item_code, prod.model)
            ctx.evidence[key].add("PRODUCTION")
            if r.daily:
                ctx.production_evidence.add(key)
            produced = sum(r.daily.values())
            schedule = norm_num(r.values.get("SCHEDULE")) or 0
            po = ProductionOrder(
                order_no=f"PR-{PERIOD_LABEL}-{ctx.stats['production_orders'] + 1:04d}",
                product_id=prod.id, section=sec.label or "",
                schedule_qty=schedule, produced_qty=produced,
                ask_till_date=norm_num(r.values.get("ASK TILL DATE")),
                completion_pct=norm_num(r.values.get("% COMP")),
                balance_qty=(schedule - produced),
                status=(ProductionStatus.completed if schedule and produced >= schedule
                        else (ProductionStatus.in_production if produced else ProductionStatus.planned)),
                report_date=dt.date(PERIOD_YEAR, PERIOD_MONTH, 31),
                source_excel=p.sheet, source_sheet=p.sheet, source_row=r.row,
                import_batch_id=ctx.batch.id, period_id=ctx.period.id,
                remarks="Excel SCHEDULE is 0 in source; balance = schedule - produced")
            ctx.db.add(po)
            ctx.db.flush()
            ctx.stats["production_orders"] += 1
            for d, q in r.daily.items():
                ctx.db.add(ProductionMovement(
                    production_order_id=po.id, quantity=q, production_date=d,
                    source_excel=p.sheet, source_row=r.row, import_batch_id=ctx.batch.id))
                ctx.stats["production_movements"] += 1
                ctx.db.add(StockMovement(
                    product_id=prod.id, movement_type=MovementType.production_output,
                    quantity=q, transaction_date=d, ref_type="production_order", ref_id=po.id,
                    source_excel=p.sheet, source_sheet=p.sheet, source_row=r.row,
                    import_batch_id=ctx.batch.id, remarks="Production output"))
                ctx.stats["stock_movements"] += 1


def import_dispatch(ctx: ImportContext, p: SheetParse):
    for sec in p.sections:
        cust = None
        if sec.label:
            status = "CONFIRMED"
            remarks = ""
            if canonical_customer(sec.label) == "TCL-2-3":
                status = "NEEDS_BUSINESS_CONFIRMATION"
                remarks = ("Historical combined TCL-2-3 section; TCL-2/TCL-3 split "
                           "not determinable from source")
            if canonical_customer(sec.label) == "OTHER":
                status = "NEEDS_BUSINESS_CONFIRMATION"
                remarks = ("Ambiguous section; extra labels: " + "; ".join(sec.extra_labels)) \
                    if sec.extra_labels else "Ambiguous section"
            cust = ctx.customer(sec.label, status=status, remarks=remarks,
                                source_sheet=p.sheet, source_row=sec.subtotal_row)
        sp = ctx.salesperson(sec.salesperson) if sec.salesperson else None
        # one OEM order per section, lines per product row
        order = None
        order_rows = [r for r in sec.rows]
        if order_rows:
            order = SalesOrder(
                order_no=f"SO-OEM-{PERIOD_LABEL}-{ctx.stats['sales_orders'] + 1:04d}",
                customer_id=cust.id if cust else None,
                order_type=OrderType.oem, period_id=ctx.period.id,
                salesperson_id=sp.id if sp else None,
                order_date=dt.date(PERIOD_YEAR, PERIOD_MONTH, 1),
                status=OrderStatus.confirmed,
                source_excel=p.sheet, source_sheet=p.sheet,
                source_row=sec.subtotal_row, import_batch_id=ctx.batch.id)
            ctx.db.add(order)
            ctx.db.flush()
            ctx.stats["sales_orders"] += 1
        for r in order_rows:
            code = r.code
            model = norm_text(r.values.get("MODEL"))
            if not model and not code:
                continue
            prod = ctx.product(code, model, category=ProductCategory.finished,
                               source_sheet=p.sheet, source_row=r.row)
            key = (prod.item_code, prod.model)
            ctx.evidence[key].add("DISPATCH")
            schedule = norm_num(r.values.get("SCHEDULE")) or 0
            dispatched = sum(r.daily.values())
            if order is not None:
                order.lines.append(SalesOrderLine(
                    product_id=prod.id, description=prod.model, quantity=schedule,
                    source_row=r.row, import_batch_id=ctx.batch.id))
                ctx.stats["sales_order_lines"] += 1
            disp = Dispatch(
                dispatch_no=f"DP-{PERIOD_LABEL}-{ctx.stats['dispatches'] + 1:04d}",
                customer_id=cust.id if cust else None,
                sales_order_id=order.id if order else None,
                salesperson_id=sp.id if sp else None,
                sales_person=sec.salesperson or "",
                period_id=ctx.period.id,
                schedule_qty=schedule, dispatched_qty=dispatched,
                ask_till_date=norm_num(r.values.get("ASK TILL DATE")),
                completion_pct=norm_num(r.values.get("% COMP")),
                balance_qty=schedule - dispatched,   # negative = over-dispatch (valid)
                status=(DispatchStatus.completed if schedule and dispatched >= schedule
                        else (DispatchStatus.partial if dispatched else DispatchStatus.pending)),
                report_date=dt.date(PERIOD_YEAR, PERIOD_MONTH, 31),
                source_excel=p.sheet, source_sheet=p.sheet, source_row=r.row,
                import_batch_id=ctx.batch.id,
                remarks="Excel OPNING STOCK is a planning metric (=BALANCE/BalanceDays); not imported as inventory")
            ctx.db.add(disp)
            ctx.db.flush()
            ctx.stats["dispatches"] += 1
            for d, q in r.daily.items():
                ctx.db.add(DispatchLine(
                    dispatch_id=disp.id, product_id=prod.id, description=prod.model,
                    quantity=q, dispatch_date=d, source_excel=p.sheet,
                    source_row=r.row, import_batch_id=ctx.batch.id))
                ctx.stats["dispatch_lines"] += 1
                ctx.db.add(StockMovement(
                    product_id=prod.id, movement_type=MovementType.dispatch,
                    quantity=q, transaction_date=d, ref_type="dispatch", ref_id=disp.id,
                    source_excel=p.sheet, source_sheet=p.sheet, source_row=r.row,
                    import_batch_id=ctx.batch.id, remarks="Dispatch"))
                ctx.stats["stock_movements"] += 1


# ---------------------------------------------------------------------------
# ORDER / PLANE sheets
# ---------------------------------------------------------------------------

def import_order_sheet(ctx: ImportContext, ws_v):
    """Local customer order log -> SalesOrder(order_type=LOCAL)."""
    rows = list(ws_v.iter_rows(values_only=True))
    for i, row in enumerate(rows, start=1):
        if i <= 2:
            continue
        if not any(row):
            continue
        odate, model, qty, cust_name, remark = (list(row) + [None] * 5)[:5]
        model, cust_name = norm_text(model), norm_text(cust_name)
        if not model or not cust_name:
            continue
        cust = ctx.customer(cust_name, source_sheet="ORDER", source_row=i)
        prod = ctx.product("", re.split(r"[\t\n]", model)[0].strip(),
                           category=ProductCategory.finished,
                           source_sheet="ORDER", source_row=i)
        d = odate.date() if isinstance(odate, dt.datetime) else dt.date(PERIOD_YEAR, PERIOD_MONTH, 1)
        status = OrderStatus.completed if norm_text(remark).upper() in SALES_STATUSES_DONE else OrderStatus.new
        order = SalesOrder(
            order_no=f"SO-LOC-{PERIOD_LABEL}-{ctx.stats['sales_orders'] + 1:04d}",
            customer_id=cust.id if cust else None,
            order_type=OrderType.local, period_id=ctx.period.id, order_date=d,
            status=status, remarks=f"Source remark: {norm_text(remark)}",
            source_excel="ORDER", source_sheet="ORDER", source_row=i,
            import_batch_id=ctx.batch.id)
        order.lines.append(SalesOrderLine(
            product_id=prod.id, description=model, quantity=norm_num(qty) or 0,
            source_row=i, import_batch_id=ctx.batch.id))
        ctx.db.add(order)
        ctx.stats["sales_orders"] += 1
        ctx.stats["sales_order_lines"] += 1


def import_plane_sheet(ctx: ImportContext, ws_v):
    """DISPATCH & PRODUCTION PLANE -> Plan (production + dispatch)."""
    rows = list(ws_v.iter_rows(values_only=True))
    plan_date_prod = ws_v["D1"].value
    plan_date_disp = ws_v["K1"].value
    plan_date_prod = plan_date_prod.date() if isinstance(plan_date_prod, dt.datetime) else dt.date(PERIOD_YEAR, PERIOD_MONTH, 14)
    plan_date_disp = plan_date_disp.date() if isinstance(plan_date_disp, dt.datetime) else dt.date(PERIOD_YEAR, PERIOD_MONTH, 14)
    for i, row in enumerate(rows, start=1):
        if i <= 2 or not any(row):
            continue
        r = list(row) + [None] * 19
        # production block A-E (0-4)
        owner, model, qty, cust, remark = norm_text(r[0]), norm_text(r[1]), norm_num(r[2]), norm_text(r[3]), norm_text(r[4])
        if model or cust:
            c = ctx.customer(cust, source_sheet="DISPATCH & PRODUCTION PLANE", source_row=i) if cust else None
            if owner:
                ctx.salesperson(owner)
            p = ctx.product("", model, category=ProductCategory.finished,
                            source_sheet="DISPATCH & PRODUCTION PLANE", source_row=i) if model else None
            ctx.db.add(Plan(
                plan_type=PlanType.production, model=model,
                product_id=p.id if p else None,
                customer_id=c.id if c else None, quantity=qty, owner=owner,
                status=remark or "PENDING", plan_date=plan_date_prod,
                source_excel="DISPATCH & PRODUCTION PLANE",
                source_sheet="DISPATCH & PRODUCTION PLANE", source_row=i,
                import_batch_id=ctx.batch.id, remarks=remark))
            ctx.stats["plans"] += 1
        # dispatch block G-L (6-11)
        model2, qty2, cust2, remark2 = norm_text(r[6]), r[7], norm_text(r[10]), norm_text(r[11])
        if model2 or cust2:
            q = norm_num(qty2)
            qty_note = "" if q is not None else norm_text(qty2)
            if qty_note:
                ctx.warnings.append(f"PLANE dispatch row {i}: non-numeric qty {qty2!r} kept as remark")
            c = ctx.customer(cust2, source_sheet="DISPATCH & PRODUCTION PLANE", source_row=i) if cust2 else None
            p = ctx.product("", model2, category=ProductCategory.finished,
                            source_sheet="DISPATCH & PRODUCTION PLANE", source_row=i) if model2 else None
            ctx.db.add(Plan(
                plan_type=PlanType.dispatch, model=model2,
                product_id=p.id if p else None,
                customer_id=c.id if c else None, quantity=q, owner="",
                status=remark2 or "RFD", plan_date=plan_date_disp,
                source_excel="DISPATCH & PRODUCTION PLANE",
                source_sheet="DISPATCH & PRODUCTION PLANE", source_row=i,
                import_batch_id=ctx.batch.id,
                remarks=(f"raw qty={qty2!r} " if qty_note else "") + (f"unit note: {qty_note}" if qty_note else "")))
            ctx.stats["plans"] += 1


# ---------------------------------------------------------------------------
# sourcing classification (after all sheets imported)
# ---------------------------------------------------------------------------

def classify_products(ctx: ImportContext):
    for key, prod in ctx.products.items():
        if prod.category == ProductCategory.raw_material:
            continue
        fam = prod.family
        trading = key in ctx.trading_evidence
        produced = key in ctx.production_evidence
        if fam and trading:
            prod.source_type = ProductSourceType.mixed
            prod.sourcing_note = (f"Family {fam} is business-confirmed MANUFACTURED; "
                                  "item also has trading transactions (STORE PO NO)")
        elif fam:
            prod.source_type = ProductSourceType.manufactured
            prod.sourcing_note = f"Business-confirmed manufactured family: {fam}"
        elif produced:
            prod.source_type = ProductSourceType.manufactured
            prod.sourcing_note = "Production output evidence in PRODUCTION sheet"
        elif "STORE" in ctx.evidence.get(key, set()):
            prod.source_type = ProductSourceType.trading
            prod.sourcing_note = "Appears only in STORE (trading tracker)"
        else:
            prod.source_type = ProductSourceType.unknown
            prod.sourcing_note = "Insufficient evidence"
        ctx.stats[f"source_type_{prod.source_type.value.lower()}"] += 1
    # leading-zero alias suggestions (no silent merge)
    codes = {k[0]: k for k in ctx.products if k[0]}
    for code, key in codes.items():
        stripped = code.lstrip("0")
        if code != stripped and stripped in codes:
            ctx.db.add(ProductAlias(
                product_id=ctx.products[codes[stripped]].id,
                alias_code=code, alias_model=key[1],
                match_rule="LEADING_ZERO_VARIANT", status="SUGGESTED",
                source_sheet=ctx.products[key].source_sheet or "",
                source_row=ctx.products[key].source_row))
            ctx.aliases_created += 1


# ---------------------------------------------------------------------------
# orchestration
# ---------------------------------------------------------------------------

SHEET_SPECS = [
    # (sheet, code_col, label_col, qty_col, day_start_col, model_as_data, importer)
    ("RAW  MATERIAL", 1, 2, 3, 9, False, "raw"),
    ("STORE", 2, 3, 4, 10, True, "store"),
    ("PRODUCTION", 1, 2, 3, 9, True, "production"),
    ("DISPATCH", 1, 2, 3, 9, False, "dispatch"),
]


def run_migration_v2(db, xlsx_path: Path, *, strict: bool = True) -> dict:
    """Parse + import the workbook into the given (staging) session."""
    xlsx_path = Path(xlsx_path)
    wbf = openpyxl.load_workbook(xlsx_path, data_only=False)
    wbv = openpyxl.load_workbook(xlsx_path, data_only=True)

    batch = ImportBatch(source_file=str(xlsx_path), period_label=PERIOD_LABEL, status="IMPORTED")
    db.add(batch)
    db.flush()
    period = ReportingPeriod(
        year=PERIOD_YEAR, month=PERIOD_MONTH, working_days=26,
        source_file=str(xlsx_path),
        notes="Confirmed period Aug-2026; workbook date columns show Aug-2025 (stale) "
              "and header cells show 2026 dates — remapped to confirmed period")
    db.add(period)
    db.flush()

    ctx = ImportContext(db, batch, period, str(xlsx_path))

    # TCL-2 / TCL-3 are separate customers going forward (no historical rows)
    ctx.customer("TCL-2", source_sheet="(confirmed business decision)")
    ctx.customer("TCL-3", source_sheet="(confirmed business decision)")

    parsed: dict[str, SheetParse] = {}
    for sheet, code_col, label_col, qty_col, day_start, mad, kind in SHEET_SPECS:
        if sheet not in wbv.sheetnames:
            ctx.warnings.append(f"sheet {sheet!r} not found")
            continue
        p = parse_report_sheet(wbf[sheet], wbv[sheet], sheet_name=sheet,
                               header_marker="ITEM CODE", code_col=code_col,
                               label_col=label_col, qty_col=qty_col,
                               day_start_col=day_start, model_as_data=mad,
                               period_year=PERIOD_YEAR, period_month=PERIOD_MONTH)
        parsed[sheet] = p
        if any(dc.stale for dc in p.day_cols):
            ctx.warnings.append(
                f"{sheet}: daily date headers are stale (Aug-2025); remapped to Aug-2026 "
                f"by day-of-month; source values preserved in raw refs")
        if kind == "raw":
            import_raw(ctx, p)
        elif kind == "store":
            import_store(ctx, p)
        elif kind == "production":
            import_production(ctx, p)
        elif kind == "dispatch":
            import_dispatch(ctx, p)
        if p.grand_total_label:
            ctx.warnings.append(
                f"{sheet}: grand-total row labeled {p.grand_total_label!r} skipped (never a customer/product)")

    # TRADING: empty template -> skip with note
    if "TRADING" in wbv.sheetnames:
        ctx.warnings.append("TRADING sheet present but empty (template); skipped")
    if "Sheet3" in wbv.sheetnames:
        ctx.warnings.append("Sheet3 (hidden scratch/pivot sheet) skipped")

    if "ORDER" in wbv.sheetnames:
        import_order_sheet(ctx, wbv["ORDER"])
    if "DISPATCH & PRODUCTION PLANE" in wbv.sheetnames:
        import_plane_sheet(ctx, wbv["DISPATCH & PRODUCTION PLANE"])

    classify_products(ctx)
    ctx.stats["aliases"] = ctx.aliases_created

    db.commit()
    return {"batch": batch, "period": period, "ctx": ctx, "parsed": parsed}
