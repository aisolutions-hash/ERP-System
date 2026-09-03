"""Sheet-aware Excel parser (v2) for the Kalika monthly workbook.

Fixes the Phase-1 defects:
* Section sub-total labels close a section -> associate *preceding* data rows
  (backward association), not the following ones.
* Grand-total rows (additive formulas right under the header) and sub-total
  rows are never treated as data.
* Daily date columns are detected from real datetime headers / serials /
  parseable strings, so every non-zero daily cell becomes a dated transaction.
* Identifiers are coerced to TEXT (integral floats -> int string); strings
  keep their whitespace-normalized form. Leading zeros in strings preserved.

The parser returns plain normalized dicts; migration_v2 maps them to models.
"""
from __future__ import annotations

import datetime as _dt
import re
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Any

import openpyxl

# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

_WS_RE = re.compile(r"\s+")


def norm_text(v: Any) -> str:
    """Normalize a text-ish cell value: collapse whitespace/newlines/tabs."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():  # numeric identifier in float form
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return _WS_RE.sub(" ", str(v).replace("\t", " ").replace("\n", " ")).strip()


def norm_num(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip())
    except (TypeError, ValueError):
        return None


# ---------------------------------------------------------------------------
# row classification
# ---------------------------------------------------------------------------

ROW_EMPTY = "EMPTY"
ROW_DATA = "DATA"
ROW_SUBTOTAL = "SECTION_SUBTOTAL"     # '=SUM(...)' aggregating a contiguous block
ROW_GRAND = "GRAND_TOTAL"             # additive formula ('=A+B+C...' refs to subtotals)
ROW_LABEL = "SECTION_LABEL"           # label-only row inside a section (forward header)

_TOTAL_LABELS = {"SUB-TOTAL", "SUBTOTAL", "G. TOTAL", "G.TOTAL", "TOTAL",
                 "GRAND TOTAL", "GRAND-TOTAL", "G. TOTAL "}


def _is_formula(cell) -> bool:
    return cell is not None and cell.data_type == "f" and isinstance(cell.value, str)


def _formula(cell) -> str:
    return cell.value if _is_formula(cell) else ""


def classify_row(ws_f, ws_v, row: int, *, code_col: int, label_col: int,
                 qty_col: int, header_row: int, model_as_data: bool) -> str:
    """Classify a row using the formula sheet (ws_f) and value sheet (ws_v).

    code_col: item-code column (STORE=B, others=A).
    label_col: section-label column (STORE=C, others=B).
    qty_col: schedule/qty column checked for aggregation formulas.
    model_as_data: when True (PRODUCTION/STORE), a row with model text but no
    code and no numeric payload is still product data (zero-activity
    placeholder), not a section label.
    """
    f_qty = ws_f.cell(row, qty_col)
    formula = _formula(f_qty)

    # grand total: additive formula directly under the header (rows hr+1..hr+2)
    if row <= header_row + 2 and formula.startswith("=") and "+" in formula:
        return ROW_GRAND
    # section subtotal: contiguous block aggregation
    if formula.startswith("=SUM("):
        return ROW_SUBTOTAL

    code = norm_text(ws_v.cell(row, code_col).value)
    label = norm_text(ws_v.cell(row, label_col).value)
    model_col = 3 if code_col == 2 else 2  # STORE model=C, others model=B
    model = norm_text(ws_v.cell(row, model_col).value)

    numeric_payload = any(
        norm_num(ws_v.cell(row, c).value) not in (None, 0)
        for c in range(1, ws_v.max_column + 1)
    )
    if code or numeric_payload:
        return ROW_DATA
    if model_as_data and model:
        return ROW_DATA
    if label or model:
        return ROW_LABEL
    return ROW_EMPTY


# ---------------------------------------------------------------------------
# daily date columns
# ---------------------------------------------------------------------------

@dataclass
class DayCol:
    col: int
    raw_value: Any
    date: _dt.date | None
    stale: bool = False   # header says a different year/month than confirmed period


def detect_day_cols(ws_v, header_row: int, start_col: int,
                    period_year: int, period_month: int) -> list[DayCol]:
    """Detect daily date columns from the header row."""
    cols: list[DayCol] = []
    for c in range(start_col, ws_v.max_column + 1):
        v = ws_v.cell(header_row, c).value
        d = None
        stale = False
        if isinstance(v, _dt.datetime):
            d = v.date()
        elif isinstance(v, _dt.date):
            d = v
        elif isinstance(v, (int, float)) and 20000 < v < 80000:
            try:
                from openpyxl.utils.datetime import from_excel
                d = from_excel(v).date()
            except Exception:
                d = None
        elif isinstance(v, str):
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y"):
                try:
                    d = _dt.datetime.strptime(v.strip()[:16], fmt).date()
                    break
                except ValueError:
                    pass
        if d is None:
            continue
        if (d.year, d.month) != (period_year, period_month):
            stale = True
            # map day-of-month into the confirmed period
            try:
                d = _dt.date(period_year, period_month, d.day)
            except ValueError:
                pass  # e.g. Feb 30 -> keep stale flag, skip below
        cols.append(DayCol(col=c, raw_value=v, date=d, stale=stale))
    return cols


@dataclass
class ParsedRow:
    row: int
    cls: str
    label: str = ""
    code: str = ""
    salesperson_cell: str = ""
    values: dict = field(default_factory=dict)       # header-name -> value
    daily: dict = field(default_factory=dict)        # date -> qty (non-zero)
    raw: dict = field(default_factory=dict)          # col -> raw value


@dataclass
class ParsedSection:
    label: str = ""
    subtotal_row: int | None = None
    salesperson: str = ""
    forward_labels: list = field(default_factory=list)
    extra_labels: list = field(default_factory=list)
    rows: list = field(default_factory=list)          # list[ParsedRow] (DATA)


@dataclass
class SheetParse:
    sheet: str
    header_row: int
    headers: dict = field(default_factory=dict)       # name -> col idx
    day_cols: list = field(default_factory=list)
    sections: list = field(default_factory=list)      # list[ParsedSection]
    skipped_counts: dict = field(default_factory=lambda: defaultdict(int))
    grand_total_label: str = ""


def parse_report_sheet(ws_f, ws_v, *, sheet_name: str, header_marker: str,
                       code_col: int, label_col: int, qty_col: int,
                       day_start_col: int, model_as_data: bool,
                       header_row: int | None = None,
                       period_year: int = 2026, period_month: int = 8) -> SheetParse:
    """Parse one of the grid-style report sheets (RAW/STORE/PRODUCTION/DISPATCH/TRADING)."""
    # locate header row by marker (e.g. 'ITEM CODE')
    hr = header_row
    if hr is None:
        for r in range(1, 12):
            for c in range(1, 12):
                if norm_text(ws_v.cell(r, c).value).upper() == header_marker:
                    hr = r
                    break
            if hr:
                break
    if hr is None:
        return SheetParse(sheet=sheet_name, header_row=0)

    headers: dict[str, int] = {}
    for c in range(1, qty_col + 9):
        name = norm_text(ws_v.cell(hr, c).value).upper()
        if name:
            headers.setdefault(name, c)
    day_cols = detect_day_cols(ws_v, hr, day_start_col, period_year, period_month)

    out = SheetParse(sheet=sheet_name, header_row=hr, headers=headers, day_cols=day_cols)
    current = ParsedSection()
    for r in range(hr + 1, ws_v.max_row + 1):
        cls = classify_row(ws_f, ws_v, r, code_col=code_col, label_col=label_col,
                           qty_col=qty_col, header_row=hr, model_as_data=model_as_data)
        if cls == ROW_EMPTY:
            continue
        label = norm_text(ws_v.cell(r, label_col).value)
        if cls == ROW_GRAND:
            out.grand_total_label = out.grand_total_label or label
            out.skipped_counts["grand_total"] += 1
            continue
        if cls == ROW_SUBTOTAL:
            # literal SUB-TOTAL labels must fall back to forward labels
            if label and label.upper() in _TOTAL_LABELS:
                label = ""
            sec_label = label or (current.forward_labels[0] if current.forward_labels else "")
            current.subtotal_row = r
            current.label = sec_label
            current.extra_labels = list(current.forward_labels)
            current.salesperson = norm_text(ws_v.cell(r, 1).value) if label_col != 1 else ""
            out.sections.append(current)
            out.skipped_counts["subtotal"] += 1
            current = ParsedSection()
            continue
        if cls == ROW_LABEL:
            current.forward_labels.append(label)
            out.skipped_counts["label"] += 1
            continue
        # DATA row
        pr = ParsedRow(row=r, cls=cls, label=label)
        pr.raw = {c: ws_v.cell(r, c).value for c in range(1, ws_v.max_column + 1)}
        for name, c in headers.items():
            pr.values[name] = ws_v.cell(r, c).value
        pr.code = norm_text(ws_v.cell(r, code_col).value)
        pr.salesperson_cell = norm_text(ws_v.cell(r, 1).value) if code_col != 1 else ""
        for dc in day_cols:
            v = norm_num(ws_v.cell(r, dc.col).value)
            if v not in (None, 0) and dc.date is not None:
                pr.daily[dc.date] = v
                if dc.stale:
                    pr.raw[f"stale_{dc.col}"] = dc.raw_value
        current.rows.append(pr)
        out.skipped_counts["data"] += 1
    # trailing rows after the last subtotal (e.g. STORE tail invoices)
    if current.rows or current.forward_labels:
        current.label = current.forward_labels[0] if current.forward_labels else ""
        out.sections.append(current)
    return out
