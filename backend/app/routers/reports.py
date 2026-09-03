"""Reports & export endpoints (CSV / Excel / PDF)."""
import csv
import io
from typing import Annotated

from fastapi import APIRouter, Depends, Query
from fastapi.responses import Response, StreamingResponse
from sqlalchemy import select, func
from sqlalchemy.orm import Session

from ..auth import CurrentUser
from ..config import settings
from ..database import get_db
from ..models import (
    Customer, Dispatch, DispatchLine, Inventory, Plant, Product, ProductionOrder,
    PurchaseOrder, RawMaterialBalance, SalesOrder, SalesOrderLine, StockMovement, Supplier,
)
from datetime import date

router = APIRouter(prefix="/reports", tags=["reports"])

EXPORT_DIR = settings.report_dir
EXPORT_DIR.mkdir(parents=True, exist_ok=True)


def _csv_response(headers, rows, filename):
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    w.writerows(rows)
    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode("utf-8-sig")),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/inventory/csv")
def inventory_csv(db: Annotated[Session, Depends(get_db)], _: CurrentUser):
    rows = db.scalars(select(Inventory)).all()
    headers = ["Product", "Item Code", "Category", "Plant", "Opening", "Received", "Issued", "Current Stock", "Min Level", "Status"]
    data = []
    for i in rows:
        status = "OK"
        if i.min_level is not None and i.current_stock <= 0:
            status = "OUT_OF_STOCK"
        elif i.min_level is not None and i.current_stock < i.min_level:
            status = "LOW"
        data.append([i.product.model if i.product else "", i.product.item_code if i.product else "",
                     i.product.category.value if i.product else "", i.plant.name if i.plant else "Main Store",
                     i.opening_stock, i.received_qty, i.issued_qty, i.current_stock, i.min_level or "", status])
    return _csv_response(headers, data, "inventory.csv")


@router.get("/movements/csv")
def movements_csv(db: Annotated[Session, Depends(get_db)], _: CurrentUser):
    rows = db.scalars(select(StockMovement).order_by(StockMovement.transaction_date.desc())).all()
    headers = ["Date", "Product", "Type", "Qty", "Plant", "Ref", "Remarks"]
    data = [[m.transaction_date, m.product.model if m.product else "", m.movement_type.value, m.quantity,
             m.plant.name if m.plant else "", m.ref_type or "", m.remarks or ""] for m in rows]
    return _csv_response(headers, data, "stock_movements.csv")


@router.get("/raw-materials/csv")
def raw_materials_csv(db: Annotated[Session, Depends(get_db)], _: CurrentUser):
    rows = db.scalars(select(RawMaterialBalance).order_by(RawMaterialBalance.report_date)).all()
    headers = ["Date", "Material", "Item Code", "Schedule", "Ask Till", "Inward", "% Comp", "Balance", "Opening"]
    data = [[b.report_date, b.product.model if b.product else "", b.product.item_code if b.product else "",
             b.schedule_qty, b.ask_till_date, b.inward_qty, b.completion_pct, b.balance_qty, b.opening_stock] for b in rows]
    return _csv_response(headers, data, "raw_materials.csv")


@router.get("/production/csv")
def production_csv(db: Annotated[Session, Depends(get_db)], _: CurrentUser):
    rows = db.scalars(select(ProductionOrder).order_by(ProductionOrder.report_date.desc())).all()
    headers = ["Order No", "Product", "Section", "Schedule", "Produced", "% Comp", "Balance", "Status", "Report Date"]
    data = [[o.order_no, o.product.model if o.product else "", o.section or "", o.schedule_qty, o.produced_qty,
             o.completion_pct, o.balance_qty, o.status.value, o.report_date] for o in rows]
    return _csv_response(headers, data, "production.csv")


@router.get("/dispatch/csv")
def dispatch_csv(db: Annotated[Session, Depends(get_db)], _: CurrentUser):
    rows = db.scalars(select(Dispatch).order_by(Dispatch.report_date.desc())).all()
    headers = ["Dispatch No", "Customer", "Plant", "Sales Person", "Schedule", "Dispatched", "% Comp", "Balance",
               "Status", "Report Date"]
    data = [[d.dispatch_no, d.customer.name if d.customer else "", d.plant.name if d.plant else "",
             d.sales_person or "", d.schedule_qty, d.dispatched_qty, d.completion_pct, d.balance_qty,
             d.status.value, d.report_date] for d in rows]
    return _csv_response(headers, data, "dispatch.csv")


@router.get("/orders/csv")
def orders_csv(db: Annotated[Session, Depends(get_db)], _: CurrentUser):
    rows = db.scalars(select(SalesOrder).order_by(SalesOrder.order_date.desc())).all()
    headers = ["Order No", "Customer", "Order Date", "Required Date", "Status", "Total Value", "Remarks"]
    data = [[o.order_no, o.customer.name if o.customer else "", o.order_date, o.required_delivery_date or "",
             o.status.value, o.total_value, o.remarks or ""] for o in rows]
    return _csv_response(headers, data, "orders.csv")


@router.get("/purchases/csv")
def purchases_csv(db: Annotated[Session, Depends(get_db)], _: CurrentUser):
    rows = db.scalars(select(PurchaseOrder).order_by(PurchaseOrder.order_date.desc())).all()
    headers = ["PO No", "Supplier", "Order Date", "Status", "Total Amount", "Notes"]
    data = [[p.po_number, p.supplier.name if p.supplier else "", p.order_date, p.status.value,
             p.total_amount, p.notes or ""] for p in rows]
    return _csv_response(headers, data, "purchases.csv")


@router.get("/customers/csv")
def customers_csv(db: Annotated[Session, Depends(get_db)], _: CurrentUser):
    rows = db.scalars(select(Customer).order_by(Customer.name)).all()
    headers = ["Name", "Code", "Plant?", "Contact", "Email", "Address"]
    data = [[c.name, c.code or "", "Yes" if c.is_plant else "No", c.contact_person or "",
             c.email or "", c.address or ""] for c in rows]
    return _csv_response(headers, data, "customers.csv")


@router.get("/suppliers/csv")
def suppliers_csv(db: Annotated[Session, Depends(get_db)], _: CurrentUser):
    rows = db.scalars(select(Supplier).order_by(Supplier.name)).all()
    headers = ["Name", "Code", "Contact", "Email", "Address"]
    data = [[s.name, s.code or "", s.contact or "", s.email or "", s.address or ""] for s in rows]
    return _csv_response(headers, data, "suppliers.csv")


@router.get("/products/csv")
def products_csv(db: Annotated[Session, Depends(get_db)], _: CurrentUser):
    rows = db.scalars(select(Product).order_by(Product.category, Product.model)).all()
    headers = ["Item Code", "Model", "Name", "Category", "UOM"]
    data = [[p.item_code or "", p.model or "", p.name or "", p.category.value, p.uom or ""] for p in rows]
    return _csv_response(headers, data, "products.csv")


def _excel_response(fname, sheet, headers, rows):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet[:31]
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.get("/excel")
def excel_report(db: Annotated[Session, Depends(get_db)], _: CurrentUser):
    """Combined Excel workbook with all modules."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_font = Font(color="FFFFFF", bold=True)

    def add_sheet(name, headers, rows):
        ws = wb.create_sheet(name[:31])
        ws.append(headers)
        for c in ws[1]:
            c.fill = hdr_fill
            c.font = hdr_font
        for r in rows:
            ws.append(r)
        for i, _ in enumerate(headers, 1):
            ws.column_dimensions[chr(64 + i)].width = 18
        return ws

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    inv = db.scalars(select(Inventory)).all()
    add_sheet("Inventory", ["Product", "Item Code", "Category", "Plant", "Opening", "Received", "Issued",
                            "Current Stock", "Min Level", "Status"],
              [[i.product.model if i.product else "", i.product.item_code if i.product else "",
                i.product.category.value if i.product else "", i.plant.name if i.plant else "Main Store",
                i.opening_stock, i.received_qty, i.issued_qty, i.current_stock, i.min_level or ""] for i in inv])

    prod = db.scalars(select(ProductionOrder).order_by(ProductionOrder.report_date)).all()
    add_sheet("Production", ["Order No", "Product", "Section", "Schedule", "Produced", "% Comp", "Balance",
                             "Status", "Report Date"],
              [[o.order_no, o.product.model if o.product else "", o.section or "", o.schedule_qty,
                o.produced_qty, o.completion_pct, o.balance_qty, o.status.value, o.report_date] for o in prod])

    disp = db.scalars(select(Dispatch).order_by(Dispatch.report_date)).all()
    add_sheet("Dispatch", ["Dispatch No", "Customer", "Plant", "Sales Person", "Schedule", "Dispatched",
                           "% Comp", "Balance", "Status", "Report Date"],
              [[d.dispatch_no, d.customer.name if d.customer else "", d.plant.name if d.plant else "",
                d.sales_person or "", d.schedule_qty, d.dispatched_qty, d.completion_pct, d.balance_qty,
                d.status.value, d.report_date] for d in disp])

    rm = db.scalars(select(RawMaterialBalance).order_by(RawMaterialBalance.report_date)).all()
    add_sheet("Raw Materials", ["Date", "Material", "Item Code", "Schedule", "Ask Till", "Inward", "% Comp",
                                "Balance", "Opening"],
              [[b.report_date, b.product.model if b.product else "", b.product.item_code if b.product else "",
                b.schedule_qty, b.ask_till_date, b.inward_qty, b.completion_pct, b.balance_qty,
                b.opening_stock] for b in rm])

    ords = db.scalars(select(SalesOrder).order_by(SalesOrder.order_date)).all()
    add_sheet("Orders", ["Order No", "Customer", "Order Date", "Required Date", "Status", "Total Value"],
              [[o.order_no, o.customer.name if o.customer else "", o.order_date, o.required_delivery_date or "",
                o.status.value, o.total_value] for o in ords])

    bos = db.scalars(select(PurchaseOrder).order_by(PurchaseOrder.order_date)).all()
    add_sheet("Purchases", ["PO No", "Supplier", "Order Date", "Status", "Total Amount"],
              [[p.po_number, p.supplier.name if p.supplier else "", p.order_date, p.status.value,
                p.total_amount] for p in bos])

    cust = db.scalars(select(Customer).order_by(Customer.name)).all()
    add_sheet("Customers", ["Name", "Code", "Plant?", "Contact", "Email", "Address"],
              [[c.name, c.code or "", "Yes" if c.is_plant else "No", c.contact_person or "",
                c.email or "", c.address or ""] for c in cust])

    supp = db.scalars(select(Supplier).order_by(Supplier.name)).all()
    add_sheet("Suppliers", ["Name", "Code", "Contact", "Email", "Address"],
              [[s.name, s.code or "", s.contact or "", s.email or "", s.address or ""] for s in supp])

    plants = db.scalars(select(Plant).order_by(Plant.name)).all()
    add_sheet("Plants", ["Code", "Name"],
              [[p.code or "", p.name] for p in plants])

    prods = db.scalars(select(Product).order_by(Product.category, Product.model)).all()
    add_sheet("Products", ["Item Code", "Model", "Name", "Category", "UOM"],
              [[p.item_code or "", p.model or "", p.name or "", p.category.value, p.uom or ""] for p in prods])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"kalika_report_{date.today().isoformat()}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ---------------------------------------------------------------------------
# Delivery Report (Phase 6K): Completed / Partially Dispatched / Not Dispatched
# ---------------------------------------------------------------------------
def _delivery_rows(db: Session, customer_id: int | None = None,
                   date_from: str = "", date_to: str = ""):
    """Per order-line delivery ledger with rollup-ready fields."""
    stmt = (select(SalesOrder, SalesOrderLine)
            .join(SalesOrderLine, SalesOrderLine.order_id == SalesOrder.id))
    if customer_id:
        stmt = stmt.where(SalesOrder.customer_id == customer_id)
    if date_from:
        stmt = stmt.where(SalesOrder.order_date >= date.fromisoformat(date_from))
    if date_to:
        stmt = stmt.where(SalesOrder.order_date <= date.fromisoformat(date_to))
    rows = db.execute(stmt).all()
    items = []
    for o, ln in rows:
        dispatched = db.scalar(
            select(func.coalesce(func.sum(DispatchLine.quantity), 0))
            .select_from(Dispatch)
            .join(DispatchLine, DispatchLine.dispatch_id == Dispatch.id)
            .where(Dispatch.sales_order_id == o.id)
        ) or 0.0
        ordered = float(ln.quantity or 0)
        disp = float(dispatched or 0)
        if ordered == 0:
            dstatus = "Not Dispatched"
        elif disp >= ordered:
            dstatus = "Completed"
        elif disp > 0:
            dstatus = "Partially Dispatched"
        else:
            dstatus = "Not Dispatched"
        items.append({
            "order_id": o.id, "order_no": o.order_no, "order_type": o.order_type.value,
            "customer_id": o.customer_id, "customer": o.customer.name if o.customer else None,
            "order_date": o.order_date.isoformat(), "period": f"{o.order_date.year}-{o.order_date.month:02d}",
            "customer_po_no": ln.customer_po_no or o.customer_po_no or "",
            "product_id": ln.product_id, "model": ln.product.model if ln.product else None,
            "item_code": ln.product.item_code if ln.product else (ln.description or ""),
            "ordered_qty": ordered, "dispatched_qty": disp,
            "balance_qty": round(ordered - disp, 4),
            "delivery_status": dstatus,
        })
    return items


@router.get("/delivery")
def delivery_report(
    db: Annotated[Session, Depends(get_db)],
    _: CurrentUser,
    customer_id: int | None = None,
    date_from: str = "",
    date_to: str = "",
    status: str = "",
):
    """Delivery report: classify each order line as Completed / Partially
    Dispatched / Not Dispatched, plus a per-customer rollup."""
    items = _delivery_rows(db, customer_id=customer_id, date_from=date_from, date_to=date_to)
    if status:
        items = [i for i in items if i["delivery_status"] == status]

    # per-customer rollup
    by = {}
    for i in items:
        key = i["customer_id"]
        e = by.setdefault(key, {
            "customer_id": key, "customer": i["customer"],
            "orders": set(), "ordered": 0.0, "dispatched": 0.0, "balance": 0.0,
            "completed": 0, "partial": 0, "not_dispatched": 0,
        })
        e["orders"].add(i["order_id"])
        e["ordered"] += i["ordered_qty"]
        e["dispatched"] += i["dispatched_qty"]
        e["balance"] += i["balance_qty"]
        if i["delivery_status"] == "Completed":
            e["completed"] += 1
        elif i["delivery_status"] == "Partially Dispatched":
            e["partial"] += 1
        else:
            e["not_dispatched"] += 1
    summary = []
    for e in by.values():
        e["ordered"] = round(e["ordered"], 4)
        e["dispatched"] = round(e["dispatched"], 4)
        e["balance"] = round(e["balance"], 4)
        e["order_count"] = len(e["orders"])
        e["delivery_pct"] = round(e["dispatched"] / e["ordered"], 4) if e["ordered"] else 0.0
        summary.append({k: (len(v) if isinstance(v, set) else v) for k, v in e.items()})
    summary.sort(key=lambda x: -(x.get("ordered") or 0))

    total_ordered = round(sum(x["ordered_qty"] for x in items), 4)
    total_dispatched = round(sum(x["dispatched_qty"] for x in items), 4)
    status_count = {}
    for i in items:
        status_count[i["delivery_status"]] = status_count.get(i["delivery_status"], 0) + 1
    return {
        "items": items,
        "summary": summary,
        "total": len(items),
        "totals": {
            "ordered": total_ordered,
            "dispatched": total_dispatched,
            "balance": round(total_ordered - total_dispatched, 4),
            "by_status": status_count,
        },
    }


@router.get("/delivery/csv")
def delivery_csv(
    db: Annotated[Session, Depends(get_db)],
    _: CurrentUser,
    customer_id: int | None = None,
    date_from: str = "",
    date_to: str = "",
):
    items = _delivery_rows(db, customer_id=customer_id, date_from=date_from, date_to=date_to)
    headers = ["Order No", "Customer", "Order Date", "PO No", "Product", "Item Code",
               "Ordered Qty", "Dispatched Qty", "Balance", "Delivery Status"]
    data = [[i["order_no"], i["customer"] or "", i["order_date"], i["customer_po_no"],
             i["model"] or "", i["item_code"], i["ordered_qty"], i["dispatched_qty"],
             i["balance_qty"], i["delivery_status"]] for i in items]
    return _csv_response(headers, data, "delivery_report.csv")